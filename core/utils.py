"""
Core utilities for PDF and Excel generation.
"""
import io
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ─── PDF Generation ──────────────────────────────────────────────────────────

def generate_pdf_response(template_name, context, filename='document.pdf'):
    """Render a template to PDF and return as HttpResponse."""
    html_string = render_to_string(template_name, context)
    html = HTML(string=html_string)
    pdf_bytes = html.write_pdf()
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def render_to_pdf_bytes(template_name, context):
    """Render a template to PDF bytes (for attachment/storage)."""
    html_string = render_to_string(template_name, context)
    html = HTML(string=html_string)
    return html.write_pdf()


# ─── Excel Generation ─────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
SUBHEADER_FILL = PatternFill(start_color='2D6A9F', end_color='2D6A9F', fill_type='solid')
SUBHEADER_FONT = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)
EVEN_ROW_FILL = PatternFill(start_color='F5F8FF', end_color='F5F8FF', fill_type='solid')


def create_excel_workbook(title='Report'):
    """Create a styled Excel workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    return wb, ws


def style_header_row(ws, row, columns):
    """Apply header styling to a row."""
    for col_num, header in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_num, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER


def style_data_row(ws, row, values, is_even=False):
    """Apply data row styling."""
    for col_num, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_num, value=value)
        if is_even:
            cell.fill = EVEN_ROW_FILL
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = BORDER


def auto_fit_columns(ws):
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 4, 50)


def excel_response(wb, filename='report.xlsx'):
    """Return Excel workbook as HttpResponse."""
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─── Pagination helper ────────────────────────────────────────────────────────

def paginate_queryset(queryset, request, per_page=20):
    """Simple pagination helper."""
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    paginator = Paginator(queryset, per_page)
    page = request.GET.get('page', 1)
    try:
        return paginator.page(page)
    except PageNotAnInteger:
        return paginator.page(1)
    except EmptyPage:
        return paginator.page(paginator.num_pages)
